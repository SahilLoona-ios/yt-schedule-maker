# ============================================================
#  excel_writer.py — Month-wise sheets, checkboxes, append mode
#
#  Sheet structure: one sheet per month ("Jan 2026", "Feb 2026"…)
#  Columns:
#   #  | Channel | Video Title | Published | Duration | Part |
#   Date | Weekday | From | To | Session | Revision | Watched | Practiced
# ============================================================

import os
import logging
from collections import defaultdict
from datetime import datetime
from typing import List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

from scheduler import WatchSession
from config import OUTPUT_FILE, WEEKDAY_BUDGET_HOURS, WEEKEND_BUDGET_HOURS

logger = logging.getLogger(__name__)

# ── Column layout ────────────────────────────────────────────
COLUMNS = [
    ("#",           4),
    ("Channel",    22),
    ("Video Title",48),
    ("Published",  12),
    ("Duration",   11),
    ("Part",        8),
    ("Date",       12),
    ("Weekday",    10),
    ("From",       10),
    ("To",         10),
    ("Session",    10),
    ("Revision",   10),
    ("Watched",    12),
    ("Practiced",  12),
]

COL_NAMES  = [c[0] for c in COLUMNS]
COL_WIDTHS = [c[1] for c in COLUMNS]

# Column indices (1-based)
COL_NUM      = 1
COL_CHANNEL  = 2
COL_TITLE    = 3
COL_PUBDATE  = 4
COL_DURATION = 5
COL_PART     = 6
COL_DATE     = 7
COL_WEEKDAY  = 8
COL_FROM     = 9
COL_TO       = 10
COL_SESSION  = 11
COL_REVISION = 12
COL_WATCHED  = 13
COL_PRACTICED= 14

TOTAL_COLS = len(COLUMNS)

# ── Color palette ────────────────────────────────────────────
HDR_COLOR       = "1A1A2E"   # Dark navy
SUBHDR_COLOR    = "16213E"
DATE_SEP_COLOR  = "0F3460"
WATCHED_COLOR   = "C8E6C9"   # Green tint when watched
PRACTICED_COLOR = "B2EBF2"   # Teal tint when practiced
DONE_BOTH_COLOR = "A5D6A7"   # Strong green when both done

VIDEO_COLORS = [
    "EBF5FB", "E8F8F0", "FEF9E7", "F4ECF7",
    "FDEDEC", "E8EAF6", "FFF3E0", "E0F7FA"
]

# Checkbox values shown in dropdown
WATCH_PENDING  = "☐ Pending"
WATCH_DONE     = "☑ Watched"
PRAC_PENDING   = "☐ Pending"
PRAC_DONE      = "☑ Practiced"


def write_excel(sessions: List[WatchSession], output_path: str = OUTPUT_FILE):
    """
    Main entry point.
    - Loads existing workbook if it exists (append mode)
    - Creates new workbook on first run
    - Groups sessions by month → writes to month-named sheets
    """
    if os.path.exists(output_path):
        wb = load_workbook(output_path)
        logger.info(f"Appending to existing workbook: {output_path}")
        mode = "append"
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]    # Remove default empty sheet
        logger.info(f"Creating new workbook: {output_path}")
        mode = "new"

    # Group sessions by "Mon YYYY"
    month_groups: dict = defaultdict(list)
    for s in sessions:
        month_key = datetime.strptime(s.session_date, "%Y-%m-%d").strftime("%b %Y")
        month_groups[month_key].append(s)

    # Sort months chronologically
    sorted_months = sorted(
        month_groups.keys(),
        key=lambda m: datetime.strptime(m, "%b %Y")
    )

    for month in sorted_months:
        month_sessions = month_groups[month]

        if month in wb.sheetnames:
            ws = wb[month]
            _append_to_sheet(ws, month_sessions)
        else:
            ws = wb.create_sheet(month)
            _init_sheet(ws)
            _append_to_sheet(ws, month_sessions)

    # Ensure sheets are sorted by date
    wb._sheets.sort(key=lambda ws: _sheet_sort_key(ws.title))

    wb.save(output_path)
    logger.info(f"Workbook saved → {output_path}  ({len(sessions)} new session(s) written)")


# ── Sheet initialisation (headers, freeze, DV, formatting) ──

def _init_sheet(ws):
    """Write header row and configure sheet-level settings."""
    # --- Header row ---
    ws.append(COL_NAMES)
    hdr_row = ws.max_row

    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(hdr_row, col_idx)
        cell.value = col_name
        cell.fill  = PatternFill("solid", fgColor=HDR_COLOR)
        cell.font  = Font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    ws.row_dimensions[hdr_row].height = 24
    ws.freeze_panes = f"A{hdr_row + 1}"

    # --- Conditional formatting (applied to data rows) ---
    data_range = f"A{hdr_row+1}:{get_column_letter(TOTAL_COLS)}9999"
    watched_col = get_column_letter(COL_WATCHED)
    prac_col    = get_column_letter(COL_PRACTICED)

    # Both done → strong green
    ws.conditional_formatting.add(data_range, FormulaRule(
        formula=[f'AND(${watched_col}{hdr_row+1}="{WATCH_DONE}",'
                 f'${prac_col}{hdr_row+1}="{PRAC_DONE}")'],
        fill=PatternFill("solid", fgColor=DONE_BOTH_COLOR)
    ))
    # Only watched → light green
    ws.conditional_formatting.add(data_range, FormulaRule(
        formula=[f'${watched_col}{hdr_row+1}="{WATCH_DONE}"'],
        fill=PatternFill("solid", fgColor=WATCHED_COLOR)
    ))
    # Only practiced → light teal
    ws.conditional_formatting.add(data_range, FormulaRule(
        formula=[f'${prac_col}{hdr_row+1}="{PRAC_DONE}"'],
        fill=PatternFill("solid", fgColor=PRACTICED_COLOR)
    ))

    ws._header_row = hdr_row    # store for use in append


def _append_to_sheet(ws, sessions: List[WatchSession]):
    """Append new session rows to an existing (or just-init'd) sheet."""
    if not sessions:
        return

    # Find header row (row 1 assumed to be header)
    header_row = 1

    # Track row numbering continuation
    existing_data_rows = ws.max_row - header_row
    row_counter        = existing_data_rows + 1

    # --- Data Validations for Watched / Practiced columns ---
    dv_watched   = DataValidation(
        type="list",
        formula1=f'"{WATCH_PENDING},{WATCH_DONE}"',
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        error="Select from dropdown",
        errorTitle="Invalid value"
    )
    dv_practiced = DataValidation(
        type="list",
        formula1=f'"{PRAC_PENDING},{PRAC_DONE}"',
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        error="Select from dropdown",
        errorTitle="Invalid value"
    )
    ws.add_data_validation(dv_watched)
    ws.add_data_validation(dv_practiced)

    # Track color per video URL
    color_map: dict = {}
    color_idx       = 0
    prev_date       = None

    for s in sessions:
        # ── Date separator row ───────────────────────────────
        if s.session_date != prev_date:
            ws.append([""] * TOTAL_COLS)
            sep_row = ws.max_row
            d       = datetime.strptime(s.session_date, "%Y-%m-%d")
            budget  = f"{WEEKEND_BUDGET_HOURS}h" if d.weekday() >= 5 else f"{WEEKDAY_BUDGET_HOURS}h"
            ws.cell(sep_row, 1).value = (
                f"📅  {d.strftime('%A, %d %B %Y')}   |   Budget: {budget}"
            )
            ws.cell(sep_row, 1).fill      = PatternFill("solid", fgColor=DATE_SEP_COLOR)
            ws.cell(sep_row, 1).font      = Font(bold=True, color="FFFFFF", size=10)
            ws.cell(sep_row, 1).alignment = Alignment(vertical="center")
            ws.merge_cells(
                start_row=sep_row, start_column=1,
                end_row=sep_row,   end_column=TOTAL_COLS
            )
            ws.row_dimensions[sep_row].height = 18
            prev_date = s.session_date

        # ── Assign color per video ───────────────────────────
        if s.url not in color_map:
            color_map[s.url] = VIDEO_COLORS[color_idx % len(VIDEO_COLORS)]
            color_idx += 1
        row_color = color_map[s.url]

        # ── Write data row ───────────────────────────────────
        d        = datetime.strptime(s.session_date, "%Y-%m-%d")
        weekday  = d.strftime("%A")
        part_str = f"{s.day_number}/{s.total_sessions}" if s.total_sessions > 1 else "Full"

        row_data = [
            row_counter,
            s.channel_name,
            s.video_title,
            s.published_at,
            s.total_duration_str,
            part_str,
            s.session_date,
            weekday,
            s.watch_from,
            s.watch_to,
            s.session_duration_str,
            s.revision_str,
            WATCH_PENDING,
            PRAC_PENDING if s.has_revision else "",
        ]
        ws.append(row_data)
        data_row = ws.max_row
        row_counter += 1

        # ── Style each cell ──────────────────────────────────
        for col_idx in range(1, TOTAL_COLS + 1):
            cell = ws.cell(data_row, col_idx)
            cell.fill      = PatternFill("solid", fgColor=row_color)
            cell.font      = Font(size=10)
            cell.alignment = Alignment(
                vertical="center",
                horizontal="center" if col_idx not in (COL_CHANNEL, COL_TITLE) else "left",
                wrap_text=(col_idx == COL_TITLE)
            )
            cell.border = _thin_border()

        # Title cell → clickable hyperlink
        title_cell           = ws.cell(data_row, COL_TITLE)
        title_cell.hyperlink = s.url
        title_cell.font      = Font(
            size=10, color="1155CC", underline="single",
            bold=(s.day_number == 1)
        )

        # Watched / Practiced cells → register in data validation
        watched_cell   = ws.cell(data_row, COL_WATCHED)
        practiced_cell = ws.cell(data_row, COL_PRACTICED)
        dv_watched.add(watched_cell)
        if s.has_revision:
            dv_practiced.add(practiced_cell)

        ws.row_dimensions[data_row].height = 18


# ── Helpers ──────────────────────────────────────────────────

def _thin_border():
    side = Side(style="thin", color="DDDDDD")
    return Border(left=side, right=side, top=side, bottom=side)


def _sheet_sort_key(title: str) -> int:
    """Converts 'Jan 2026' → sortable integer like 202601."""
    try:
        dt = datetime.strptime(title, "%b %Y")
        return dt.year * 100 + dt.month
    except ValueError:
        return 999999    # Non-month sheets go to end
